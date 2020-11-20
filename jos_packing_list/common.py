# encoding=utf-8
import requests
import time
import json
import settings
from datetime import datetime

from sqlalchemy.orm import sessionmaker
from models import *
from sqlalchemy import select, func, union_all, cast, Unicode, null, case

from random import randint

engine = create_engine('postgresql+psycopg2://%s:%s@%s/%s' % (settings.USER, settings.PASSWORD,
                                                              settings.DATA_HOST, settings.DATABASE), echo=True, client_encoding='utf8')
Session = sessionmaker(bind=engine)

def get_params(request):
    params = request.params
    body = request.body.read()
    if body:
        try:
            rec = json.loads(body)
        except Exception as e:
            s = body.split('=')
            rec = {s[0]: s[1]}
        for item in rec:
            params[item] = rec.get(item)
    return params


def get_store_name(url, token, id):
    headers = {"Token": token}
    r = requests.get('%s/api/stores/%s' % (url, id), headers=headers)
    if r.status_code != 200:
        raise Exception(r.text)
    return r.json().get("name")


def get_sale_order(url, token, id):
    headers = {"Token": token}
    r = requests.get('%s/api/sale_orders/%s' %(url, id), headers=headers)
    sale_order = json.loads(r.text, object_pairs_hook=ObjectDict)
    if not sale_order or sale_order.status != 6:
        return False
    lines = get_sale_order_lines(url, token, id, sale_order.partner_id)
    total_quantity = 0
    for item in lines:
        total_quantity = total_quantity + (item.quantity or 0)
    sale_order.total_quantity = total_quantity
    sale_order.sale_order_lines = lines
    return sale_order


def get_sale_order_lines(url, token, id, partner_id):
    headers = {"Token": token}
    r = requests.get('%s/api/sale_orders/%s/sale_order_lines' %(url, id), headers=headers)
    results = []
    product_ids = set([])
    p_ids = []
    lines = json.loads(r.text, object_pairs_hook=ObjectDict)
    for line in lines:
        product_ids.add(str(line.product_id))
    _product_ids = ",".join(product_ids)
    product_map = get_product_map(url, token, _product_ids, partner_id)
    for item in lines:
        if item.product_id in p_ids:
            for re in results:
                if re.product_id == item.product_id:
                    re.quantity = re.quantity + item.quantity
        else:
            p_ids.append(item.product_id)
            pro = product_map.get(item.product_id)
            if not pro:
                raise Exception(u'商品 %s 对应的合作伙伴物料号不存在' % item.product_name)
            item.product_name = pro.get("name")
            item.product_code = pro.get("code")
            results.append(item)
    return results

def get_product_map(url, token, ids, partner_id):
    headers = {"Token": token}
    params = {"product_id__in": ids, "partner_id": partner_id}
    product_map = {}
    r = requests.get('%s/api/partner_products' %url, headers=headers, params=params)
    for item in r.json().get("root"):
        product_map[item.get("product_id")] = item
    return product_map


def to_code(id):
    date = datetime.now().strftime("%Y-%m-%d")
    code = "PL-"+date+"-%05d" % id
    return code

def to_pdf(data):
    from xhtml2pdf import pisa
    from cStringIO import StringIO

    io = StringIO()

    pisa.CreatePDF(data, io)
    v = io.getvalue()
    io.close()
    return v

def get_result(id, session, set_of_book):
    packing_list = session.query(PackingList).filter(PackingList.id == id, PackingList.set_of_book == set_of_book).first()
    results = packing_list.to_dict()
    results.packing_list_lines = []
    packing_list_lines = session.query(PackingListLine).filter(PackingListLine.packing_list_id == packing_list.id, PackingListLine.set_of_book == set_of_book).all()
    for p in packing_list_lines:
        packing_list_line = p.to_dict()
        packing_list_line.product_lines = []
        product_lines = session.query(ProductLine).filter(ProductLine.packing_list_line_id == p.id, ProductLine.set_of_book == set_of_book).all()
        for pr in product_lines:
            product_line = pr.to_dict()
            packing_list_line.product_lines.append(product_line)
        results.packing_list_lines.append(packing_list_line)
    return results


def get_one_page(session, id, packing_list_line_id, set_of_book):
    packing_list = session.query(PackingList).filter(PackingList.id == id, PackingList.set_of_book == set_of_book).first()
    result = packing_list.to_dict()
    packing_list_line = session.query(PackingListLine).filter(PackingListLine.id == packing_list_line_id, PackingListLine.packing_list_id == id, PackingListLine.set_of_book == set_of_book).first()
    packing_list_line = packing_list_line.to_dict()
    packing_list_line.product_lines = []
    product_lines = session.query(ProductLine).filter(ProductLine.packing_list_line_id == packing_list_line_id, ProductLine.set_of_book == set_of_book).all()
    for pr in product_lines:
        packing_list_line.product_lines.append(pr.to_dict())
    result.packing_list_line = packing_list_line
    return result


def packing_reports(session, params, set_of_book):
    format = params.get("format")
    if format and format == "excel":
        start_date = params.get("start_date")
        end_date = params.get("end_date")
        form_code = params.get("form_code") or None
        code = params.get("code") or None
        start = params.get("start") or None
        limit = params.get("limit") or None
        partner_name = params.get("partner_name") or None
        fields1 = [packing_lists.c.partner_name, packing_lists.c.store_name, packing_lists.c.order_id, packing_lists.c.des, packing_lists.c.store, packing_lists.c.created_at, packing_lists.c.created_user, packing_lists.c.executed_at, packing_lists.c.executed_user, packing_lists.c.form_code, packing_lists.c.box_count, packing_lists.c.total_quantity, packing_list_lines.c.code, packing_list_lines.c.total_quantity.label("line_total_quantity"), product_lines.c.product_name, product_lines.c.packing_quantity, product_lines.c.product_code]
        q1 = select(fields1, from_obj=(packing_lists.join(packing_list_lines, packing_list_lines.c.packing_list_id == packing_lists.c.id).join(product_lines, product_lines.c.packing_list_line_id == packing_list_lines.c.id))).where(packing_lists.c.status == 6).where(packing_lists.c.set_of_book == set_of_book).where(product_lines.c.packing_quantity != 0)
        if start_date:
            q1 = q1.where(packing_lists.c.executed_at >= start_date)
        if end_date:
            q1 = q1.where(packing_lists.c.executed_at <= end_date)
        if code:
            try:
                code = int(code)
                q1 = q1.where(packing_lists.c.id == code)
            except:
                q1 = q1.where(packing_lists.c.code == code)
        if form_code:
            try:
                form_code = int(form_code)
                q1 = q1.where(packing_lists.c.form_id == form_code)
            except:
                q1 = q1.where(packing_lists.c.form_code == form_code)
        if partner_name:
            q1 = q1.where(packing_lists.c.partner_name == partner_name)
        result2 = session.execute(q1).fetchall()
        _packing_lists = ObjectDict()
        for item in result2:
            if item.form_code not in _packing_lists:
                _packing_lists[item.form_code] = ObjectDict(partner_name=item.partner_name, store_name=item.store_name, order_id=item.order_id, des=item.des, store=item.store, created_at=item.created_at.strftime("%Y-%m-%d %H:%M:%S"), created_user=item.created_user, executed_at=item.executed_at.strftime("%Y-%m-%d %H:%M:%S"), executed_user=item.executed_user, box_count=item.box_count, total_quantity=item.total_quantity)
                _packing_lists[item.form_code].packing_list_lines = ObjectDict({item.code:ObjectDict(total_quantity=item.line_total_quantity, product_lines=ObjectDict({item.product_name: ObjectDict(packing_quantity=item.packing_quantity, product_code=item.product_code)}))})
            else:
                if item.code in _packing_lists[item.form_code].packing_list_lines:
                    if item.product_name in _packing_lists[item.form_code].packing_list_lines[item.code].product_lines:
                        pass
                    else:
                        _packing_lists[item.form_code].packing_list_lines[item.code].product_lines[item.product_name] = ObjectDict(packing_quantity=item.packing_quantity, product_code=item.product_code)
                else:
                    _packing_lists[item.form_code].packing_list_lines[item.code] = ObjectDict(total_quantity=item.line_total_quantity, product_lines=ObjectDict({item.product_name: ObjectDict(packing_quantity=item.packing_quantity, product_code=item.product_code)}))
        file_name = u"装箱单报表.xlsx"
        from openpyxl.workbook import Workbook

        wb = Workbook(write_only=True)
        sheet = wb.create_sheet()
        sheet.title = u'装箱单报表'

        # create table header
        headers = [u"商业伙伴", u"仓库", u"单号", u"总数量", u"创建时间", u"完成时间", u"总箱数", u"装箱单号", u"装箱总数", u"商品名称", u"装箱数量", u"商品编码"]
        sheet.append(headers)
        for item1 in _packing_lists:
            p = _packing_lists.get(item1)
            values = [p.partner_name, p.store_name, item1, p.total_quantity, p.created_at, p.executed_at, p.box_count, "", "", "", "", ""]
            sheet.append(values)
            for item2 in p.packing_list_lines:
                pll = p.packing_list_lines.get(item2)
                values = ["", "", "", "", "", "", "", item2, pll.total_quantity, "", "", ""]
                sheet.append(values)
                for item3 in pll.product_lines:
                    pl = pll.product_lines.get(item3)
                    values = ["", "", "", "", "", "", "", "", "", item3, pl.packing_quantity, pl.product_code]
                    sheet.append(values)
        from cStringIO import StringIO
        io = StringIO()
        wb.save(io)
        return {'file_name': file_name, 'body': io.getvalue()}
    else:
        start_date = params.get("start_date")
        end_date = params.get("end_date")
        form_code = params.get("form_code") or None
        code = params.get("code") or None
        start = params.get("start") or None
        limit = params.get("limit") or None
        partner_name = params.get("partner_name") or None
        fields = [packing_lists.c.partner_name, packing_lists.c.order_id, packing_lists.c.store, packing_lists.c.des, packing_lists.c.created_at, packing_lists.c.created_user, packing_lists.c.executed_at, packing_lists.c.executed_user, packing_lists.c.form_code, packing_lists.c.box_count, packing_lists.c.total_quantity, packing_list_lines.c.code]
        q = select(fields, from_obj=(packing_lists.join(packing_list_lines, packing_list_lines.c.packing_list_id == packing_lists.c.id))).where(packing_lists.c.status == 6).where(packing_lists.c.set_of_book == set_of_book)
        if start_date:
            q = q.where(packing_lists.c.executed_at >= start_date)
        if end_date:
            q = q.where(packing_lists.c.executed_at <= end_date)
        if partner_name:
            q = q.where(packing_lists.c.partner_name == partner_name)
        if code:
            try:
                code = int(code)
                q = q.where(packing_lists.c.id == code)
            except:
                q = q.where(packing_lists.c.code == code)
        if form_code:
            try:
                form_code = int(form_code)
                q = q.where(packing_lists.c.form_id == form_code)
            except:
                q = q.where(packing_lists.c.form_code == form_code)
        if start is not None and limit:
            x = q.alias("x")
            total = session.execute(x.count()).scalar()
            _results = session.execute(q.offset(start).limit(limit).order_by(packing_lists.c.executed_at.desc())).fetchall()
        else:
            _results = session.execute(q.order_by(packing_lists.c.executed_at.desc())).fetchall()
            total = len(_results)
        results = []
        for item in _results:
            k = ObjectDict()
            k.partner_name = item.partner_name
            k.order_id = item.order_id
            k.des = item.des
            k.created_at = item.created_at.strftime("%Y-%m-%d %H:%M:%S")
            k.created_user = item.created_user
            k.executed_at = item.executed_at.strftime("%Y-%m-%d %H:%M:%S")
            k.executed_user = item.executed_user
            k.form_code = item.form_code
            k.code = item.code
            k.box_count = item.box_count
            k.total_quantity = item.total_quantity
            k.store = item.store
            results.append(k)
        return {"total": total, "result": results}


def packing_report_lines(session, params, set_of_book):
    code = params.get("code")
    format = params.get("format")
    fields = [packing_list_lines.c.code, product_lines.c.product_name, product_lines.c.packing_quantity]
    q = select(fields, from_obj=(packing_lists.join(packing_list_lines, packing_list_lines.c.packing_list_id == packing_lists.c.id).join(product_lines, product_lines.c.packing_list_line_id == packing_list_lines.c.id))).where(packing_lists.c.status == 6).where(product_lines.c.packing_quantity != 0)
    if code:
        q = q.where(packing_list_lines.c.code == code)
    _results = session.execute(q.order_by(packing_lists.c.created_at.desc())).fetchall()
    total = len(_results)
    results = []
    for item in _results:
        k = ObjectDict()
        k.code = item.code
        k.product_name = item.product_name
        k.packing_quantity = item.packing_quantity
        results.append(k)
    if format and format == "excel":
        file_name = u"装箱单明细报表.xlsx"
        from openpyxl.workbook import Workbook

        wb = Workbook(write_only=True)
        sheet = wb.create_sheet()
        sheet.title = u'装箱单报表明细'

        # create table header
        headers = [u"序号", u"装箱单号", u"商品名称", u"装箱数量"]
        sheet.append(headers)
        for index, item in enumerate(results):
            values = [index+1, item.code, item.product_name, item.packing_quantity]
            sheet.append(values)
        from cStringIO import StringIO
        io = StringIO()
        wb.save(io)
        return {'file_name': file_name, 'body': io.getvalue()}
    else:
        return {"total": total, "results": results}


def check_quantity(id, session, set_of_book):
    data = get_result(id, session, set_of_book).packing_list_lines
    products1 = {}
    products2 = {}
    for p in data[0].product_lines:
        products1[p.product_id] = p.quantity
    for item in data:
        for p in item.product_lines:
            if p.product_id in products2:
                products2[p.product_id] += p.packing_quantity
            else:
                products2[p.product_id] = p.packing_quantity
    for x in products1:
        if products1.get(x) != products2.get(x):
            return False
    return True

def check_empty_box(id, session, set_of_book):
    packing_list_lines = get_result(id, session, set_of_book).packing_list_lines
    for packing_list_line in packing_list_lines:
        state = 0
        for product_line in packing_list_line.product_lines:
            if product_line.packing_quantity > 0:
                state = 1
                break
        if state == 0:
            return False
    return True


def delete_packing_list_line(id, session, set_of_book):
    packing_list_line = session.query(PackingListLine).filter(PackingListLine.id == id, PackingListLine.set_of_book == set_of_book).first()
    product_lines = session.query(ProductLine).filter(ProductLine.packing_list_line_id == id, ProductLine.set_of_book == set_of_book).all()
    for p in product_lines:
        session.delete(p)
    session.flush()
    session.delete(packing_list_line)
    session.flush()

def transfer_weight(weight):
    if type(weight) in (float, int):
        return weight
    elif type(weight) == str:
        weight = weight.strip()
        return float(weight.split(' ')[0])